#openpyxl支持刪除Excel文件中的工作表，但Pandas不支持直接刪除
from openpyxl import load_workbook

#載入現有的Excel文件
workbook = load_workbook('example.xlsx')

#列出所有工作表名稱
print("原有工作表：", workbook.sheetnames)

#刪除特定工作表（例如名為'Sheet1'的工作表）
if 'Sheet1' in workbook.sheetnames:
    workbook.remove(workbook['Sheet1'])

#保存修改後的工作簿
workbook.save('example_modified_openpyxl.xlsx')

#列出修改後的工作表名稱
print("修改後估做表：", workbook.sheetnames)