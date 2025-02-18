from openpyxl import load_workbook

#載入現有的Excel文件
workbook = load_workbook('example.xlsx')
sheet = workbook.active

#修改特定單元的數據
sheet['B2'] = 27.00     #將Widget A的價格從25.5改為27.00
#可以這麼寫是因為表格簡單，可以一眼就看到要修改的位置是B2，但如果哪天表格欄位對調，就不能這樣寫了
#要透過層層回區勸尋找

#尋找'Widget A'的位置並修其價格
for row in sheet.iter_rows(min_row = 2):    #從第二行開始，跳過標題行
    if row[0].value == 'Widget A':          #假設第一列是產品名稱
        row[1].value = 27.00                #修改價格
        break                               #找到後跳出迴圈
#在不知道特定值位置，openpyxl要遍歷整個工作表

#保存修改後的工作簿
workbook.save('example_modified_openpyxl.xlsx')